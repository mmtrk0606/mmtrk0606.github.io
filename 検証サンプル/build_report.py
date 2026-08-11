# -*- coding: utf-8 -*-
"""
ポートフォリオ用 不具合レポートサンプル
架空の会議室予約システム「Rezerva」を検証した想定で作成
※実在しないサービスです
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

OUT = "out/不具合レポート_Rezerva_サンプル.xlsx"

# ── 配色 ──────────────────────────────────────────────
NAVY = "1F3864"
NAVY_L = "D9E2F3"
GRAY_L = "F2F2F2"
RED, RED_BG = "C00000", "FCE4E4"
ORG, ORG_BG = "C55A11", "FDF0E6"
GRY, GRY_BG = "595959", "F2F2F2"
GREEN = "2E7D32"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

F_TITLE = Font(name="游ゴシック", size=18, bold=True, color=NAVY)
F_H = Font(name="游ゴシック", size=10.5, bold=True, color="FFFFFF")
F_B = Font(name="游ゴシック", size=10)
F_BB = Font(name="游ゴシック", size=10, bold=True)
F_S = Font(name="游ゴシック", size=9, color="595959")

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_W = Alignment(horizontal="center", vertical="center", wrap_text=True)

SEV = {
    "高": (RED, RED_BG),
    "中": (ORG, ORG_BG),
    "低": (GRY, GRY_BG),
}

# ── 不具合データ ───────────────────────────────────────
BUGS = [
    dict(
        sev="高", cat="入力チェック", screen="新規予約",
        title="終了時刻が開始時刻より前でも予約が登録できる",
        steps="1. 一般利用者でログインする\n"
              "2. [新規予約] を開く\n"
              "3. 会議室「会議室A」、利用日「2026-08-18」を選択する\n"
              "4. 開始時刻「16:00」、終了時刻「14:00」を選択する\n"
              "5. 件名を入力し、[この内容で予約する] を押す",
        expect="「終了時刻は開始時刻より後に設定してください」という趣旨のエラーが表示され、登録されない。",
        actual="エラーは表示されず、予約が登録される。登録後の一覧には利用時間が「16:00 – 14:00」と表示される。",
        repro="100%（5回中5回）",
        env="macOS 15.5 / Chrome 139 / 1440px",
        cause="開始・終了の前後関係を検証する処理が見当たりません。"
              "登録時のバリデーションに、終了 > 開始 の判定を追加する必要があります。",
        shot="bug-01",
    ),
    dict(
        sev="高", cat="業務ロジック", screen="新規予約 / 予約一覧",
        title="時間帯が一部重なる予約を、同じ会議室に登録できる",
        steps="1. 会議室Aに「10:00 – 12:00」の予約が存在する状態にする\n"
              "2. 別の利用者でログインする\n"
              "3. 会議室A・同日で「11:00 – 13:00」を指定して予約する",
        expect="「指定の時間帯はすでに予約されています」と表示され、登録されない。",
        actual="登録できてしまう。予約一覧に同じ会議室の重複した予約が2件並ぶ。\n"
               "なお「10:00 – 12:00」と完全に同一の時間帯を指定した場合はエラーになる。",
        repro="100%（5回中5回）",
        env="macOS 15.5 / Chrome 139 / 1440px",
        cause="重複判定が開始時刻・終了時刻の完全一致で行われている可能性があります。"
              "「既存の開始 < 新規の終了 かつ 新規の開始 < 既存の終了」で判定する必要があります。",
        shot="bug-02",
    ),
    dict(
        sev="高", cat="権限", screen="予約の編集",
        title="他の利用者の予約を、URLを直接指定することで編集・削除できる",
        steps="1. 一般利用者Aでログインする\n"
              "2. 他の利用者Bの予約IDを確認する（一覧の[詳細]リンクから取得できる）\n"
              "3. ブラウザのアドレス欄に /reservations/{利用者Bの予約ID}/edit を直接入力して開く",
        expect="自分以外の予約は開けず、権限がない旨のメッセージまたは一覧へのリダイレクトとなる。",
        actual="利用者Bの予約の編集画面が表示され、時間・件名の変更と削除ができてしまう。\n"
               "画面上は予約者名が「山田 一郎」と他人のまま表示される。",
        repro="100%（3回中3回）",
        env="macOS 15.5 / Chrome 139 / 1440px",
        cause="一覧の表示時には自分の予約に絞り込まれている一方、編集画面の表示・更新処理で"
              "「予約者 = ログイン中の利用者」の確認が行われていないと考えられます。",
        shot="bug-03",
    ),
    dict(
        sev="中", cat="業務ロジック", screen="自分の予約",
        title="キャンセル期限（開始1時間前）を過ぎてもキャンセルできる",
        steps="1. 開始時刻まで1時間を切っている予約がある状態にする\n"
              "   （例：現在 09:35 / 予約 10:00 開始）\n"
              "2. [自分の予約] を開く\n"
              "3. 該当予約の [キャンセル] を押す",
        expect="仕様どおりであれば、ボタンが非活性になるか、押下時に期限切れの案内が表示される。",
        actual="通常どおりキャンセルが完了し、会議室が空き枠に戻る。",
        repro="100%（3回中3回）",
        env="macOS 15.5 / Chrome 139 / 1440px",
        cause="キャンセル可否の判定処理が見当たりません。"
              "なお、この制限自体が仕様として必要かどうかは確認させてください。",
        shot="bug-04",
    ),
    dict(
        sev="中", cat="入力チェック", screen="新規予約（API）",
        title="件名の文字数制限が画面側にしかなく、APIへ直接送ると制限を超えて登録できる",
        steps="1. 画面上では件名に51文字以上を入力できないことを確認する\n"
              "2. 開発者ツールの Network から予約登録のリクエストを確認する\n"
              "3. 同じリクエストを、件名に200文字を入れて再送する",
        expect="サーバー側でも文字数を検証し、エラーが返る。",
        actual="200文字の件名で登録される。予約一覧では件名がセル幅を超えて表示が崩れる。",
        repro="100%（3回中3回）",
        env="macOS 15.5 / Chrome 139",
        cause="入力チェックが画面側のみに実装されていると考えられます。"
              "登録・更新のAPIでも同じ条件で検証する必要があります。",
        shot="—",
    ),
    dict(
        sev="中", cat="表示", screen="新規予約",
        title="営業時間外（22:00より後）の時刻が選択肢に表示される",
        steps="1. [新規予約] を開く\n"
              "2. 終了時刻のプルダウンを開き、末尾まで確認する",
        expect="営業時間は 8:00〜22:00 のため、22:00 より後の選択肢は表示されない。",
        actual="22:30 / 23:00 / 23:30 が選択でき、選択した状態で登録もできる。\n"
               "なお画面下部の注釈には「営業時間は 8:00〜22:00 です」と記載されている。",
        repro="100%（3回中3回）",
        env="macOS 15.5 / Chrome 139 / 1440px",
        cause="選択肢の生成範囲が 24:00 までになっていると考えられます。"
              "営業時間の設定値を参照して生成するか、上限を 22:00 に修正する必要があります。",
        shot="bug-06",
    ),
    dict(
        sev="中", cat="業務ロジック", screen="新規予約",
        title="同時保有できる予約の上限（5件）に、キャンセル済みの予約が含まれている",
        steps="1. 予約を5件登録する\n"
              "2. そのうち2件をキャンセルする（残り3件が有効）\n"
              "3. 新しい予約を登録する",
        expect="有効な予約は3件のため、追加で登録できる。",
        actual="「予約の上限に達しています」と表示され、登録できない。\n"
               "キャンセル済みの2件を含めて5件と数えられている。",
        repro="100%（3回中3回）",
        env="macOS 15.5 / Chrome 139 / 1440px",
        cause="件数の集計時に、状態が「キャンセル済」のレコードが除外されていないと考えられます。",
        shot="—",
    ),
    dict(
        sev="低", cat="表示", screen="予約一覧（スマートフォン）",
        title="スマートフォン表示で予約一覧の表が画面外にはみ出す",
        steps="1. 画面幅 390px（iPhone 相当）で [予約一覧] を開く\n"
              "2. 表の右端を確認する",
        expect="表が画面内に収まる。または表だけが横スクロールし、ページ全体は横に動かない。",
        actual="表が画面幅を超え、ページ全体が横スクロールする。\n"
               "「状態」「詳細」の列が初期表示では見えない。",
        repro="100%（3回中3回）",
        env="iOS 18 Safari / 画面幅 390px、Chrome のデバイスエミュレーションでも再現",
        cause="表を囲う要素に横スクロールの指定がないためと考えられます。",
        shot="bug-08",
    ),
    dict(
        sev="低", cat="表示", screen="新規予約",
        title="登録に失敗したとき、原因が分からないメッセージだけが表示される",
        steps="1. すでに予約されている時間帯を指定して登録する\n"
              "2. 画面上部のメッセージを確認する",
        expect="「指定の時間帯はすでに予約されています」など、利用者が次に何をすればよいか分かる文言が表示される。",
        actual="「エラーが発生しました。」とのみ表示される。\n"
               "どの項目に問題があるかが分からず、入力し直す箇所を特定できない。",
        repro="100%（5回中5回）",
        env="macOS 15.5 / Chrome 139 / 1440px",
        cause="例外をまとめて捕捉し、共通の文言を表示していると考えられます。",
        shot="bug-09",
    ),
    dict(
        sev="低", cat="データ整合", screen="会議室管理 / 予約一覧",
        title="会議室を削除しても、その会議室の予約が残り続ける",
        steps="1. 管理者でログインする\n"
              "2. [会議室管理] から、予約が入っている会議室を削除する\n"
              "3. [予約一覧] を開く",
        expect="削除時に「この会議室には予約が3件あります」と警告が出る。\n"
               "または、関連する予約もあわせて取り消される。",
        actual="警告なく削除でき、予約一覧には会議室名が「不明な会議室」と表示された予約が残る。\n"
               "この予約は編集もキャンセルもできない。",
        repro="100%（3回中3回）",
        env="macOS 15.5 / Chrome 139 / 1440px",
        cause="会議室の削除時に、関連する予約の有無が確認されていないと考えられます。"
              "削除を禁止するか、論理削除にするかは方針をご確認ください。",
        shot="bug-10",
    ),
]

# ── 確認した観点 ───────────────────────────────────────
VIEWS = [
    ("入力チェック", "未入力での送信", "必須項目を空のまま送信したときの挙動", "○", ""),
    ("入力チェック", "文字数の上限・下限", "件名50文字、備考500文字の境界値", "×", "No.5"),
    ("入力チェック", "不正な値", "利用日に過去日・存在しない日付を指定", "○", ""),
    ("入力チェック", "前後関係", "開始時刻と終了時刻の前後関係", "×", "No.1"),
    ("業務ロジック", "時間帯の重複", "同一会議室で時間が重なる予約の禁止", "×", "No.2"),
    ("業務ロジック", "予約可能な範囲", "営業時間 8:00〜22:00 の範囲外の指定", "×", "No.6"),
    ("業務ロジック", "予約件数の上限", "1利用者あたり5件までの制限", "×", "No.7"),
    ("業務ロジック", "キャンセル期限", "開始1時間前を過ぎた予約のキャンセル", "×", "No.4"),
    ("権限", "画面の表示制御", "一般利用者に管理メニューが表示されないか", "○", ""),
    ("権限", "他人のデータ", "URL直接指定による他人の予約への操作", "×", "No.3"),
    ("権限", "ログアウト後", "ログアウト後に戻るボタンで画面を再表示", "○", ""),
    ("データ整合", "関連データの削除", "会議室・利用者を削除したときの予約の扱い", "×", "No.10"),
    ("表示", "スマートフォン表示", "画面幅390pxでのレイアウト崩れ", "×", "No.8"),
    ("表示", "エラーメッセージ", "失敗時に原因と次の操作が分かるか", "×", "No.9"),
    ("表示", "一覧の並び順", "予約一覧が時間順に並ぶか", "○", ""),
    ("表示", "0件のとき", "予約が1件もないときの表示", "○", ""),
]


def style_header(ws, row, cols, height=30):
    ws.row_dimensions[row].height = height
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_H
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = CENTER_W
        cell.border = BORDER


# ══════════════════════════════════════════════════════
wb = Workbook()

# ── シート1：概要 ──────────────────────────────────────
ws = wb.active
ws.title = "概要"
ws.sheet_view.showGridLines = False
ws["B2"] = "不具合レポート"
ws["B2"].font = F_TITLE
ws["B3"] = "会議室予約システム「Rezerva」／ 検証結果のご報告"
ws["B3"].font = Font(name="游ゴシック", size=11, color="595959")

info = [
    ("検証対象", "会議室予約システム Rezerva（ステージング環境）"),
    ("検証期間", "2026年8月10日 〜 2026年8月11日"),
    ("検証範囲", "予約の登録・編集・キャンセル／会議室管理／権限（管理者・一般利用者）"),
    ("検証環境", "macOS 15.5 ／ Chrome 139 ／ 画面幅 1440px・390px"),
    ("検証方法", "画面からの操作による検証、および開発者ツールでのリクエスト確認"),
    ("報告件数", "10件（高 3件／中 4件／低 3件）"),
]
r = 6
for k, v in info:
    ws.cell(r, 2, k).font = F_BB
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=NAVY_L)
    ws.cell(r, 2).alignment = Alignment(vertical="center")
    ws.cell(r, 2).border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(r, 3, v).font = F_B
    ws.cell(r, 3).alignment = Alignment(vertical="center")
    for c in (3, 4):
        ws.cell(r, c).border = BORDER
    ws.row_dimensions[r].height = 24
    r += 1

# 重要度別の集計
r += 1
ws.cell(r, 2, "重要度の内訳").font = F_BB
r += 1
head = ["重要度", "件数", "対応の目安"]
for i, h in enumerate(head):
    ws.cell(r, 2 + i, h)
style_header(ws, r, 4, 24)
ws.cell(r, 1).fill = PatternFill()  # A列はヘッダー対象外
ws.cell(r, 1).border = Border()
counts = [
    ("高", 3, "リリース前の修正を推奨します。データの不整合や情報の閲覧につながります。"),
    ("中", 4, "仕様の確認のうえ、優先して修正されることを推奨します。"),
    ("低", 3, "利用に支障はありませんが、使い勝手の面で改善の余地があります。"),
]
for sev, n, note in counts:
    r += 1
    fg, bg = SEV[sev]
    ws.cell(r, 2, sev).font = Font(name="游ゴシック", size=10, bold=True, color=fg)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=bg)
    ws.cell(r, 2).alignment = CENTER
    ws.cell(r, 3, n).font = F_B
    ws.cell(r, 3).alignment = CENTER
    ws.cell(r, 4, note).font = F_B
    ws.cell(r, 4).alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22
    for c in range(2, 5):
        ws.cell(r, c).border = BORDER

r += 3
ws.cell(r, 2, "本レポートはポートフォリオ用の制作サンプルです。").font = Font(name="游ゴシック", size=9, bold=True, color=RED)
ws.cell(r + 1, 2, "Rezerva は実在しないサービスであり、記載の不具合・環境・日付はすべて架空のものです。").font = F_S
ws.cell(r + 2, 2, "実際のご依頼では、ご指定の形式（Excel／Googleスプレッドシート等）で納品いたします。").font = F_S

for col, w in [("A", 3), ("B", 16), ("C", 12), ("D", 76)]:
    ws.column_dimensions[col].width = w

# ── シート2：不具合一覧 ────────────────────────────────
ws = wb.create_sheet("不具合一覧")
ws.sheet_view.showGridLines = False
cols = [
    ("No", 6), ("重要度", 8), ("分類", 12), ("発生画面", 16),
    ("事象", 34), ("再現手順", 46), ("期待する動作", 34), ("実際の動作", 40),
    ("再現性", 12), ("発生環境", 22), ("原因と思われる箇所", 40),
    ("スクショ", 10), ("起票日", 11), ("状態", 10),
]
for i, (h, w) in enumerate(cols, start=1):
    ws.cell(1, i, h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, 1, len(cols), 32)

for i, b in enumerate(BUGS, start=1):
    r = i + 1
    fg, bg = SEV[b["sev"]]
    vals = [
        i, b["sev"], b["cat"], b["screen"], b["title"], b["steps"],
        b["expect"], b["actual"], b["repro"], b["env"], b["cause"],
        b["shot"], "2026-08-11", "未対応",
    ]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(r, c, v)
        cell.font = F_B
        cell.alignment = WRAP_TOP if c in (5, 6, 7, 8, 11) else CENTER_W
        cell.border = BORDER
    ws.cell(r, 1).alignment = CENTER
    ws.cell(r, 2).font = Font(name="游ゴシック", size=10, bold=True, color=fg)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=bg)
    ws.cell(r, 5).font = F_BB
    ws.row_dimensions[r].height = 118

ws.freeze_panes = "E2"
ws.auto_filter.ref = f"A1:N{len(BUGS) + 1}"

dv = DataValidation(type="list", formula1='"未対応,対応中,修正済,再確認済,対応不要"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"N2:N{len(BUGS) + 1}")

# ── シート3：確認した観点 ──────────────────────────────
ws = wb.create_sheet("確認した観点")
ws.sheet_view.showGridLines = False
ws["A1"] = "確認した観点の一覧"
ws["A1"].font = Font(name="游ゴシック", size=13, bold=True, color=NAVY)
ws["A2"] = "○＝問題なし　／　×＝不具合として報告（該当のNoを記載）"
ws["A2"].font = F_S

hs = [("分類", 14), ("観点", 24), ("確認した内容", 52), ("結果", 8), ("報告No", 10)]
for i, (h, w) in enumerate(hs, start=1):
    ws.cell(4, i, h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, 4, len(hs), 26)

for i, (cat, view, desc, res, no) in enumerate(VIEWS, start=5):
    for c, v in enumerate([cat, view, desc, res, no], start=1):
        cell = ws.cell(i, c, v)
        cell.font = F_B
        cell.alignment = Alignment(vertical="center", wrap_text=True) if c == 3 else CENTER
        cell.border = BORDER
    ws.cell(i, 4).font = Font(name="游ゴシック", size=10, bold=True, color=RED if res == "×" else GREEN)
    ws.row_dimensions[i].height = 24

ws.freeze_panes = "A5"

# ── 印刷設定（PDF配布時に横に切れないようにする） ──────────
for sh, paper in [("概要", Workbook().active.PAPERSIZE_A4),
                  ("不具合一覧", Workbook().active.PAPERSIZE_A3),
                  ("確認した観点", Workbook().active.PAPERSIZE_A4)]:
    s = wb[sh]
    s.page_setup.orientation = "landscape"
    s.page_setup.paperSize = paper
    s.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    s.page_setup.fitToWidth = 1
    s.page_setup.fitToHeight = 0
    s.print_options.horizontalCentered = True
    s.page_margins.left = s.page_margins.right = 0.3
    s.page_margins.top = s.page_margins.bottom = 0.4
wb["不具合一覧"].print_title_rows = "1:1"
wb["確認した観点"].print_title_rows = "4:4"

wb.save(OUT)
print(f"生成完了: {OUT}　（不具合 {len(BUGS)}件／観点 {len(VIEWS)}件）")
